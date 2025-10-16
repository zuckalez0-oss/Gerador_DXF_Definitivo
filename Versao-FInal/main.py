# main.py (VERSÃO OTIMIZADA COM LAYOUT RESPONSIVO)

from openpyxl.styles import PatternFill, Font
from openpyxl import load_workbook
import sys
import os
import json
import pandas as pd
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                             QHBoxLayout, QPushButton, QLabel, QTextEdit, 
                             QFileDialog, QProgressBar, QMessageBox, QGroupBox,
                             QFormLayout, QLineEdit, QComboBox, QTableWidget, 
                             QTableWidgetItem, QDialog, QInputDialog, QHeaderView,
                             QSplitter) # <<< IMPORTAÇÃO NECESSÁRIA >>>
from PyQt5.QtCore import Qt

# <<< IMPORTAÇÕES DAS CLASSES ENCAPSULADAS >>>
from code_manager import CodeGenerator
from history_manager import HistoryManager
from history_dialog import HistoryDialog
from processing import ProcessThread
from nesting_dialog import NestingDialog
from dxf_engine import get_dxf_bounding_box # <<< IMPORTAÇÃO NECESSÁRIA >>>
from calculo_cortes import orquestrar_planos_de_corte

# =============================================================================
# ESTILO VISUAL DA APLICAÇÃO (QSS - Qt StyleSheet)
# =============================================================================
INOVA_PROCESS_STYLE = """
/* ================================================================================
   Estilo INOVA PROCESS (v8 - Tema Preto & Branco de Alto Contraste)
================================================================================ */

/* Passo 1: Fundo principal e cor de texto padrão */
QWidget {
    background-color: #111111; /* Preto suave para o fundo principal */
    color: #FFFFFF;           /* Branco puro para o texto */
    font-family: 'Segoe UI', Arial, sans-serif;
    font-size: 7pt; 
    border: none;
}

QLabel {
    color: #FFFFFF;
    background: transparent;
}

/* Divisores (Splitter) */
QSplitter::handle { background-color: #333333; }
QSplitter::handle:hover { background-color: #FFFFFF; }
QSplitter::handle:pressed { background-color: #CCCCCC; }

/* Passo 2: Contêineres com fundo de cinza escuro para contraste */
QGroupBox, QTableWidget, QListView {
    background-color: #222222; 
    border: 1px solid #444444; /* Borda cinza */
    border-radius: 8px;
}
QGroupBox {
    margin-top: 1em; 
    font-weight: bold;
}

/* Passo 3: Cor de Destaque - Títulos com fundo branco e texto preto */
QGroupBox::title {
    subcontrol-origin: margin;
    subcontrol-position: top center;
    padding: 2px 8px;
    background-color: #FFFFFF; /* Destaque invertido (branco) */
    color: #000000;           /* Destaque invertido (preto) */
    border-radius: 4px;
    font-weight: bold;
}

/* Passo 4: Campos de Input e ComboBox */
QLineEdit, QTextEdit, QComboBox, QDoubleSpinBox, QSpinBox {
    background-color: #111111;   /* Mesmo fundo do QWidget */
    border: 1px solid #444444;   /* Borda cinza */
    border-radius: 4px;
    padding: 4px; 
    color: #FFFFFF;
}
QLineEdit:focus, QTextEdit:focus, QComboBox:focus, QDoubleSpinBox:focus, QSpinBox:focus {
    border: 1px solid #FFFFFF; /* Foco com borda branca */
}
QLineEdit::placeholder {
    color: #777777;
}

/* Detalhes do ComboBox */
QComboBox::drop-down { border: none; }
QComboBox::down-arrow {
    image: url(C:/Users/mathe/Desktop/INOVA_PROCESS/down_arrow.png); /* Manter o ícone de seta (assumindo que seja branco) */
    width: 10px; height: 10px; margin-right: 8px;
}
QComboBox QAbstractItemView {
    background-color: #222222;
    border: 1px solid #FFFFFF;
    selection-background-color: #FFFFFF;
    selection-color: #000000;
    outline: 0px;
}

/* Passo 5: Botões Padrão com tons de cinza */
QPushButton {
    background-color: #333333;   /* Cinza médio */
    color: #DDDDDD;           /* Branco suave */
    font-weight: bold;
    padding: 4px 8px; 
    border-radius: 4px;
}
QPushButton:hover { background-color: #555555; } /* Cinza mais claro no hover */
QPushButton:pressed { background-color: #222222; }

/* Botão Primário (Invertido, como os títulos) */
QPushButton#primaryButton { background-color: #FFFFFF; color: #000000; }
QPushButton#primaryButton:hover { background-color: #CCCCCC; }

/* Botões de estado (Mantidos por usabilidade) */
QPushButton#successButton { background-color: #107C10; color: #FFFFFF; }
QPushButton#successButton:hover { background-color: #159d15; }
QPushButton#warningButton { background-color: #DCA307; color: #1A202C; }
QPushButton#warningButton:hover { background-color: #f0b92a; }

/* Passo 6: Tabela */
QTableWidget {
    gridline-color: #444444;
}
QHeaderView::section {
    background-color: #222222;
    color: #DDDDDD; /* Branco suave para não competir com os dados */
    padding: 4px;
    border: 1px solid #444444;
    font-weight: bold;
}
QTableWidget::item {
    color: #FFFFFF;
    font-size: 8pt;
    padding: 4px;
}
/* Seleção da tabela também usa o destaque invertido */
QTableWidget::item:selected {
    background-color: #FFFFFF;
    color: #000000;
}

/* Barra de Log */
QTextEdit#logExecution {
    font-family: 'Courier New', Courier, monospace;
    background-color: #222222;
    color: #4FD1C5; /* Mantive o ciano aqui para um look "terminal" clássico */
}

/* Barras de Rolagem */
QScrollBar:vertical { border: none; background: #222222; width: 12px; margin: 0; }
QScrollBar::handle:vertical { background: #444444; min-height: 20px; border-radius: 6px; }
QScrollBar::handle:vertical:hover { background: #666666; }
QScrollBar:horizontal { border: none; background: #222222; height: 12px; margin: 0; }
QScrollBar::handle:horizontal { background: #444444; min-width: 20px; border-radius: 6px; }
QScrollBar::handle:horizontal:hover { background: #666666; }
QScrollBar::add-line, QScrollBar::sub-line { border: none; background: none; }
"""

# =============================================================================
# CLASSE PRINCIPAL DA INTERFACE GRÁFICA
# =============================================================================
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Gerador de Desenhos Técnicos e DXF INP - NOROACO")
        self.setGeometry(100, 100, 1280, 850) 
        self.setMinimumSize(1100, 800)

        self.code_generator = CodeGenerator()
        self.history_manager = HistoryManager()
        
        self.colunas_df = ['nome_arquivo', 'forma', 'espessura', 'qtd', 'largura', 'altura', 'diametro', 'rt_base', 'rt_height', 'trapezoid_large_base', 'trapezoid_small_base', 'trapezoid_height', 'furos']
        self.colunas_df = ['nome_arquivo', 'forma', 'espessura', 'qtd', 'largura', 'altura', 'diametro', 'rt_base', 'rt_height', 'trapezoid_large_base', 'trapezoid_small_base', 'trapezoid_height', 'furos', 'dxf_path']
        self.manual_df = pd.DataFrame(columns=self.colunas_df)
        self.excel_df = pd.DataFrame(columns=self.colunas_df)
        self.furos_atuais = []
        self.project_directory = None

        self.initUI() # Chama o método que constrói a UI
        self.connect_signals() # Chama o método que conecta os eventos
        
        self.set_initial_button_state()
        self.update_dimension_fields(self.forma_combo.currentText())

    def initUI(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)

        # --- Layout Superior (Inputs e Furos) ---
        top_h_layout = QHBoxLayout()
        
        # <<< MUDANÇA ESTRUTURAL 1: PAINEL ESQUERDO COM LARGURA MÍNIMA >>>
        left_panel_widget = QWidget()
        left_v_layout = QVBoxLayout(left_panel_widget)
        left_v_layout.setContentsMargins(0,0,0,0) # Remove margens internas
        left_panel_widget.setMinimumWidth(450) # Impede o "esmagamento"

        # --- Grupo 1: Projeto ---
        project_group = QGroupBox("1. Projeto")
        project_layout = QVBoxLayout()
        self.start_project_btn = QPushButton("Iniciar Novo Projeto...")
        self.history_btn = QPushButton("Ver Histórico de Projetos")
        project_layout.addWidget(self.start_project_btn)
        project_layout.addWidget(self.history_btn)
        project_group.setLayout(project_layout)
        left_v_layout.addWidget(project_group)
        
        # --- Grupo 2: Carregar Planilha ---
        file_group = QGroupBox("2. Carregar Planilha (Opcional)")
        file_layout = QVBoxLayout()
        self.file_label = QLabel("Nenhum projeto ativo.")
        file_button_layout = QHBoxLayout()
        self.select_file_btn = QPushButton("Selecionar Planilha")
        self.import_dxf_btn = QPushButton("Importar DXF(s)") # <<< NOVO BOTÃO >>>
        self.clear_excel_btn = QPushButton("Limpar Planilha")
        file_button_layout.addWidget(self.select_file_btn)
        file_button_layout.addWidget(self.import_dxf_btn)
        file_button_layout.addWidget(self.clear_excel_btn)
        file_layout.addWidget(self.file_label)
        file_layout.addLayout(file_button_layout)
        file_group.setLayout(file_layout)
        left_v_layout.addWidget(file_group)

        # --- Grupo 3: Informações da Peça ---
        manual_group = QGroupBox("3. Informações da Peça")
        manual_layout = QFormLayout()
        manual_layout.setLabelAlignment(Qt.AlignRight)
        manual_layout.setVerticalSpacing(8)
        self.projeto_input = QLineEdit()
        self.projeto_input.setReadOnly(True)
        manual_layout.addRow("Nº do Projeto Ativo:", self.projeto_input)
        self.nome_input = QLineEdit()
        self.generate_code_btn = QPushButton("Gerar Código")
        name_layout = QHBoxLayout()
        name_layout.addWidget(self.nome_input)
        name_layout.addWidget(self.generate_code_btn)
        name_layout.setSpacing(5)
        manual_layout.addRow("Nome/ID da Peça:", name_layout)
        self.forma_combo = QComboBox()
        self.forma_combo.addItems(['rectangle', 'circle', 'right_triangle', 'trapezoid', 'dxf_shape'])
        self.espessura_input, self.qtd_input = QLineEdit(), QLineEdit()
        manual_layout.addRow("Forma:", self.forma_combo)
        manual_layout.addRow("Espessura (mm):", self.espessura_input)
        manual_layout.addRow("Quantidade:", self.qtd_input)
        self.largura_input, self.altura_input = QLineEdit(), QLineEdit()
        self.diametro_input, self.rt_base_input, self.rt_height_input = QLineEdit(), QLineEdit(), QLineEdit()
        self.trapezoid_large_base_input, self.trapezoid_small_base_input, self.trapezoid_height_input = QLineEdit(), QLineEdit(), QLineEdit()
        self.largura_row = [QLabel("Largura:"), self.largura_input]; manual_layout.addRow(*self.largura_row)
        self.altura_row = [QLabel("Altura:"), self.altura_input]; manual_layout.addRow(*self.altura_row)
        self.diametro_row = [QLabel("Diâmetro:"), self.diametro_input]; manual_layout.addRow(*self.diametro_row)
        self.rt_base_row = [QLabel("Base Triângulo:"), self.rt_base_input]; manual_layout.addRow(*self.rt_base_row)
        self.rt_height_row = [QLabel("Altura Triângulo:"), self.rt_height_input]; manual_layout.addRow(*self.rt_height_row)
        self.trap_large_base_row = [QLabel("Base Maior:"), self.trapezoid_large_base_input]; manual_layout.addRow(*self.trap_large_base_row)
        self.trap_small_base_row = [QLabel("Base Menor:"), self.trapezoid_small_base_input]; manual_layout.addRow(*self.trap_small_base_row)
        self.trap_height_row = [QLabel("Altura:"), self.trapezoid_height_input]; manual_layout.addRow(*self.trap_height_row)
        manual_group.setLayout(manual_layout)
        left_v_layout.addWidget(manual_group)
        left_v_layout.addStretch()
        
        top_h_layout.addWidget(left_panel_widget) # Adiciona o painel esquerdo ao layout horizontal

        # --- Grupo 4: Furos (Painel Direito) ---
        furos_main_group = QGroupBox("4. Adicionar Furos")
        furos_main_layout = QVBoxLayout()
        self.rep_group = QGroupBox("Furação Rápida")
        rep_layout = QFormLayout()
        self.rep_diam_input, self.rep_offset_input = QLineEdit(), QLineEdit()
        rep_layout.addRow("Diâmetro Furos:", self.rep_diam_input)
        rep_layout.addRow("Offset Borda:", self.rep_offset_input)
        self.replicate_btn = QPushButton("Replicar Furos")
        rep_layout.addRow(self.replicate_btn)
        self.rep_group.setLayout(rep_layout)
        furos_main_layout.addWidget(self.rep_group)
        man_group = QGroupBox("Furos Manuais")
        man_layout = QVBoxLayout()
        man_form_layout = QFormLayout()
        self.diametro_furo_input, self.pos_x_input, self.pos_y_input = QLineEdit(), QLineEdit(), QLineEdit()
        man_form_layout.addRow("Diâmetro:", self.diametro_furo_input)
        man_form_layout.addRow("Posição X:", self.pos_x_input)
        man_form_layout.addRow("Posição Y:", self.pos_y_input)
        self.add_furo_btn = QPushButton("Adicionar Furo Manual")
        man_layout.addLayout(man_form_layout)
        man_layout.addWidget(self.add_furo_btn)
        self.furos_table = QTableWidget(0, 4)
        self.furos_table.setMaximumHeight(150)
        self.furos_table.setHorizontalHeaderLabels(["Diâmetro", "Pos X", "Pos Y", "Ação"])
        man_layout.addWidget(self.furos_table)
        man_group.setLayout(man_layout)
        furos_main_layout.addWidget(man_group)
        furos_main_group.setLayout(furos_main_layout)
        top_h_layout.addWidget(furos_main_group, stretch=1)

        # Container para o layout superior
        top_container_widget = QWidget()
        top_container_widget.setLayout(top_h_layout)

        # --- Grupo 5: Lista de Peças ---
        list_group = QGroupBox("5. Lista de Peças para Produção")
        list_layout = QVBoxLayout()
        self.pieces_table = QTableWidget()
        self.table_headers = [col.replace('_', ' ').title() for col in self.colunas_df] + ["Ações"]
        self.pieces_table.setColumnCount(len(self.table_headers))
        self.pieces_table.setHorizontalHeaderLabels(self.table_headers)
        self.pieces_table.verticalHeader().setDefaultSectionSize(28) 
        self.pieces_table.setMinimumHeight(120)
       
        list_layout.addWidget(self.pieces_table)
        self.dir_label = QLabel("Nenhum projeto ativo. Inicie um novo projeto.")
        self.dir_label.setStyleSheet("font-style: italic; color: grey;")
        list_layout.addWidget(self.dir_label)
        process_buttons_layout = QHBoxLayout()
        self.conclude_project_btn = QPushButton("Projeto Concluído")
        self.export_excel_btn = QPushButton("Exportar para Excel")
        self.process_pdf_btn, self.process_dxf_btn, self.process_all_btn = QPushButton("Gerar PDFs"), QPushButton("Gerar DXFs"), QPushButton("Gerar PDFs e DXFs")
        process_buttons_layout.addWidget(self.export_excel_btn)
        process_buttons_layout.addWidget(self.conclude_project_btn)
        process_buttons_layout.addStretch()
        self.calculate_nesting_btn = QPushButton("Calcular Aproveitamento")
        process_buttons_layout.addWidget(self.calculate_nesting_btn)
        process_buttons_layout.addWidget(self.process_pdf_btn)
        process_buttons_layout.addWidget(self.process_dxf_btn)
        process_buttons_layout.addWidget(self.process_all_btn)
        list_layout.addLayout(process_buttons_layout)
        list_group.setLayout(list_layout)

        # --- Barra de Log/Execução ---
        log_group = QGroupBox("Log de Execução")
        log_layout = QVBoxLayout()
        self.log_text = QTextEdit()
        self.log_text.setObjectName("logExecution") # Adicionado para estilo
        log_layout.addWidget(self.log_text)
        log_group.setLayout(log_layout)
        
        # <<< MUDANÇA ESTRUTURAL 2: USO DO QSPLITTER PARA O LAYOUT VERTICAL >>>
        v_splitter = QSplitter(Qt.Vertical)
        #v_splitter.addWidget(top_container_widget)
        v_splitter.addWidget(list_group)
        v_splitter.addWidget(log_group)

        v_splitter.setStretchFactor(0, 1)
        v_splitter.setStretchFactor(1, 0)
        v_splitter.setSizes([400, 150])

        self.add_piece_btn = QPushButton("Adicionar Peça à Lista")
        main_layout.addWidget(top_container_widget)
        main_layout.addWidget(v_splitter)
        main_layout.addWidget(self.add_piece_btn)

        # --- Barra de Progresso ---
        self.progress_bar = QProgressBar()
        main_layout.addWidget(self.progress_bar)
        
        self.statusBar().showMessage("Pronto")
        
        # --- Aplicação de Estilos Específicos via objectName ---
        self.start_project_btn.setObjectName("primaryButton")
        self.conclude_project_btn.setObjectName("successButton")
        self.calculate_nesting_btn.setObjectName("warningButton")

    def connect_signals(self):
        """Método para centralizar todas as conexões de sinais e slots."""
        self.calculate_nesting_btn.clicked.connect(self.open_nesting_dialog)
        self.start_project_btn.clicked.connect(self.start_new_project)
        self.history_btn.clicked.connect(self.show_history_dialog)
        self.select_file_btn.clicked.connect(self.select_file)
        self.import_dxf_btn.clicked.connect(self.import_dxfs) # <<< CONEXÃO DO SINAL >>>
        self.clear_excel_btn.clicked.connect(self.clear_excel_data)
        self.generate_code_btn.clicked.connect(self.generate_piece_code)
        self.add_piece_btn.clicked.connect(self.add_manual_piece)
        self.forma_combo.currentTextChanged.connect(self.update_dimension_fields)
        self.replicate_btn.clicked.connect(self.replicate_holes)
        self.add_furo_btn.clicked.connect(self.add_furo_temp)
        self.process_pdf_btn.clicked.connect(self.start_pdf_generation)
        self.process_dxf_btn.clicked.connect(self.start_dxf_generation)
        self.process_all_btn.clicked.connect(self.start_all_generation)
        self.conclude_project_btn.clicked.connect(self.conclude_project)
        self.export_excel_btn.clicked.connect(self.export_project_to_excel)

    # =====================================================================
    # O RESTANTE DAS FUNÇÕES (MÉTODOS) PERMANECE O MESMO
    # ... (Cole aqui todos os seus métodos de 'start_new_project' até 'delete_furo_temp')
    # =====================================================================
    # --- INÍCIO: NOVA FUNÇÃO PARA OFFSET DINÂMICO (COMPARTILHADA) ---
    def _get_dynamic_offset_and_margin(self, espessura, default_offset, default_margin):
        """Retorna o offset e a margem com base na espessura."""
        # --- CORREÇÃO: A função agora prioriza o input do usuário se for diferente do padrão '8'. ---
        # Se o usuário inseriu um valor diferente do padrão (8), usa o valor do usuário.
        if abs(default_offset - 8.0) > 1e-5:
            return default_offset, default_margin

        if 0 < espessura <= 6.35: return 5, 10
        elif 6.35 < espessura <= 15.88: return 10, default_margin
        elif 15.88 < espessura <= 20: return 17, default_margin
        elif abs(espessura - 22.22) < 1e-5: return 20, default_margin
        elif 25.4 <= espessura <= 38: return 25, default_margin
        return default_offset, default_margin
    # --- FIM: NOVA FUNÇÃO PARA OFFSET DINÂMICO ---

    def start_new_project(self):
        parent_dir = QFileDialog.getExistingDirectory(self, "Selecione a Pasta Principal para o Novo Projeto")
        if not parent_dir: return
        project_name, ok = QInputDialog.getText(self, "Novo Projeto", "Digite o nome ou número do novo projeto:")
        if ok and project_name:
            project_path = os.path.join(parent_dir, project_name)
            if os.path.exists(project_path):
                reply = QMessageBox.question(self, 'Diretório Existente', f"A pasta '{project_name}' já existe.\nDeseja usá-la como o diretório do projeto ativo?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
                if reply == QMessageBox.No: return
            else:
                try: os.makedirs(project_path)
                except OSError as e: QMessageBox.critical(self, "Erro ao Criar Pasta", f"Não foi possível criar o diretório do projeto:\n{e}"); return
            self._clear_session(clear_project_number=True)
            self.project_directory = project_path
            self.projeto_input.setText(project_name)
            self.dir_label.setText(f"Projeto Ativo: {self.project_directory}")
            self.dir_label.setStyleSheet("font-style: normal; color: #E2E8F0;") # Cor do texto do tema
            self.log_text.append(f"\n--- NOVO PROJETO INICIADO: {project_name} ---")
            self.log_text.append(f"Arquivos serão salvos em: {self.project_directory}")
            self.set_initial_button_state()

    def set_initial_button_state(self):
        is_project_active = self.project_directory is not None
        has_items = not (self.excel_df.empty and self.manual_df.empty)
        self.calculate_nesting_btn.setEnabled(is_project_active and has_items)
        self.start_project_btn.setEnabled(True)
        self.history_btn.setEnabled(True)
        self.select_file_btn.setEnabled(is_project_active)
        self.import_dxf_btn.setEnabled(is_project_active) # <<< ATUALIZAÇÃO DE ESTADO >>>
        self.clear_excel_btn.setEnabled(is_project_active and not self.excel_df.empty)
        self.generate_code_btn.setEnabled(is_project_active)
        self.add_piece_btn.setEnabled(is_project_active)
        self.replicate_btn.setEnabled(is_project_active)
        self.add_furo_btn.setEnabled(is_project_active)
        self.process_pdf_btn.setEnabled(is_project_active and has_items)
        self.process_dxf_btn.setEnabled(is_project_active and has_items)
        self.process_all_btn.setEnabled(is_project_active and has_items)
        self.conclude_project_btn.setEnabled(is_project_active and has_items)
        self.export_excel_btn.setEnabled(is_project_active and has_items)
        self.progress_bar.setVisible(False)

    def show_history_dialog(self):
        dialog = HistoryDialog(self.history_manager, self)
        if dialog.exec_() == QDialog.Accepted:
            loaded_pieces = dialog.loaded_project_data
            if loaded_pieces:
                project_number_loaded = loaded_pieces[0].get('project_number') if loaded_pieces and 'project_number' in loaded_pieces[0] else dialog.project_list_widget.currentItem().text()
                self.start_new_project_from_history(project_number_loaded, loaded_pieces)
    
    def start_new_project_from_history(self, project_name, pieces_data):
        parent_dir = QFileDialog.getExistingDirectory(self, f"Selecione uma pasta para o projeto '{project_name}'")
        if not parent_dir: return
        project_path = os.path.join(parent_dir, project_name)
        os.makedirs(project_path, exist_ok=True)
        self._clear_session(clear_project_number=True)
        self.project_directory = project_path
        self.projeto_input.setText(project_name)
        self.excel_df = pd.DataFrame(columns=self.colunas_df)
        self.manual_df = pd.DataFrame(pieces_data)
        self.dir_label.setText(f"Projeto Ativo: {self.project_directory}"); self.dir_label.setStyleSheet("font-style: normal; color: #E2E8F0;")
        self.log_text.append(f"\n--- PROJETO DO HISTÓRICO CARREGADO: {project_name} ---")
        self.update_table_display()
        self.set_initial_button_state()

    def start_pdf_generation(self): self.start_processing(generate_pdf=True, generate_dxf=False)
    def start_dxf_generation(self): self.start_processing(generate_pdf=False, generate_dxf=True)
    def start_all_generation(self): self.start_processing(generate_pdf=True, generate_dxf=True)

    def start_processing(self, generate_pdf, generate_dxf):
        if not self.project_directory:
            QMessageBox.warning(self, "Nenhum Projeto Ativo", "Inicie um novo projeto antes de gerar arquivos."); return
        project_number = self.projeto_input.text().strip()
        if not project_number:
            QMessageBox.warning(self, "Número do Projeto Ausente", "Por favor, defina um número para o projeto ativo."); return
        # --- CORREÇÃO FUTUREWARNING: Concatena apenas os dataframes não vazios ---
        dfs_to_concat = [df for df in [self.excel_df, self.manual_df] if not df.empty]
        if not dfs_to_concat:
            QMessageBox.warning(self, "Aviso", "A lista de peças está vazia."); return
        combined_df = pd.concat(dfs_to_concat, ignore_index=True)
        # --- FIM CORREÇÃO ---
        self.set_buttons_enabled_on_process(False)
        self.progress_bar.setVisible(True); self.progress_bar.setValue(0); self.log_text.clear()
        self.process_thread = ProcessThread(combined_df.copy(), generate_pdf, generate_dxf, self.project_directory, project_number)
        self.process_thread.update_signal.connect(self.log_text.append)
        self.process_thread.progress_signal.connect(self.progress_bar.setValue)
        self.process_thread.finished_signal.connect(self.processing_finished)
        self.process_thread.start()

    def processing_finished(self, success, message):
        self.set_buttons_enabled_on_process(True); self.progress_bar.setVisible(False)
        msgBox = QMessageBox.information if success else QMessageBox.critical
        msgBox(self, "Concluído" if success else "Erro", message); self.statusBar().showMessage("Pronto")
    
    def conclude_project(self):
        project_number = self.projeto_input.text().strip()
        if not project_number:
            QMessageBox.warning(self, "Projeto sem Número", "O projeto ativo não tem um número definido.")
            return
        reply = QMessageBox.question(self, 'Concluir Projeto', f"Deseja salvar e concluir o projeto '{project_number}'?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            # --- CORREÇÃO FUTUREWARNING: Concatena apenas os dataframes não vazios ---
            dfs_to_concat = [df for df in [self.excel_df, self.manual_df] if not df.empty]
            if dfs_to_concat:
                combined_df = pd.concat(dfs_to_concat, ignore_index=True)
            # --- FIM CORREÇÃO ---
                combined_df['project_number'] = project_number
                combined_df['project_number'] = project_number
                self.history_manager.save_project(project_number, combined_df)
                self.log_text.append(f"Projeto '{project_number}' salvo no histórico.")
            self._clear_session(clear_project_number=True)
            self.project_directory = None
            self.dir_label.setText("Nenhum projeto ativo. Inicie um novo projeto."); self.dir_label.setStyleSheet("font-style: italic; color: grey;")
            self.set_initial_button_state()
            self.log_text.append(f"\n--- PROJETO '{project_number}' CONCLUÍDO ---")

    def open_nesting_dialog(self):
        # --- CORREÇÃO FUTUREWARNING: Concatena apenas os dataframes não vazios ---
        dfs_to_concat = [df for df in [self.excel_df, self.manual_df] if not df.empty]
        if not dfs_to_concat:
            QMessageBox.warning(self, "Lista Vazia", "Não há peças na lista para calcular o aproveitamento.")
            return
        combined_df = pd.concat(dfs_to_concat, ignore_index=True)
        # --- CORREÇÃO: Inclui 'circle' na verificação de formas válidas ---
        valid_df = combined_df[combined_df['forma'].isin(['rectangle', 'circle', 'right_triangle', 'trapezoid', 'dxf_shape'])].copy()
        if valid_df.empty:
            QMessageBox.information(self, "Nenhuma Peça Válida", "O cálculo de aproveitamento só pode ser feito com peças da forma 'rectangle', 'circle', 'right_triangle', 'trapezoid' ou 'dxf_shape'.")
            return
        # Passa o DataFrame com as formas válidas para o diálogo
        dialog = NestingDialog(valid_df, self)
        dialog.exec_()

    def export_project_to_excel(self):
        chapa_largura_str, ok1 = QInputDialog.getText(self, "Parâmetro de Aproveitamento", "Largura da Chapa (mm):", text="3000")
        if not ok1: return
        chapa_altura_str, ok2 = QInputDialog.getText(self, "Parâmetro de Aproveitamento", "Altura da Chapa (mm):", text="1500")
        if not ok2: return
        offset_str, ok3 = QInputDialog.getText(self, "Parâmetro de Aproveitamento", "Offset entre Peças (mm):", text="8")
        if not ok3: return
        margin_str, ok4 = QInputDialog.getText(self, "Parâmetro de Aproveitamento", "Margem da Chapa (mm):", text="10")
        if not ok4: return
        try:
            chapa_largura = float(chapa_largura_str)
            chapa_altura = float(chapa_altura_str)
            offset = float(offset_str)
            margin = float(margin_str)
        except (ValueError, TypeError):
            QMessageBox.critical(self, "Erro de Entrada", "Valores de chapa e offset devem ser numéricos.")
            return

        project_number = self.projeto_input.text().strip()
        if not project_number:
            QMessageBox.warning(self, "Nenhum Projeto Ativo", "Inicie um novo projeto para poder exportá-lo.")
            return

        dfs_to_concat = [df for df in [self.excel_df, self.manual_df] if not df.empty]
        if not dfs_to_concat:
            QMessageBox.warning(self, "Lista Vazia", "Não há peças na lista para exportar.")
            return
        combined_df = pd.concat(dfs_to_concat, ignore_index=True)

        default_filename = os.path.join(self.project_directory, f"CUSTO_PLASMA-LASER_V4_NOVA_{project_number}.xlsx")
        save_path, _ = QFileDialog.getSaveFileName(self, "Salvar Resumo do Projeto", default_filename, "Excel Files (*.xlsx)")
        if not save_path:
            return

        self.set_buttons_enabled_on_process(False)
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.log_text.clear()
        self.log_text.append("Iniciando exportação para Excel...")
        QApplication.processEvents()

        try:
            template_path = 'CUSTO_PLASMA-LASER_V4_NOVA.xlsx'
            if not os.path.exists(template_path):
                QMessageBox.critical(self, "Template Não Encontrado", f"O arquivo modelo '{template_path}' não foi encontrado.")
                return

            wb = load_workbook(template_path)
            ws = wb.active
            self.log_text.append("Preenchendo lista de peças...")
            QApplication.processEvents()

            start_row = 1
            while ws.cell(row=start_row, column=1).value is not None: start_row += 1

            for index, (_, row_data) in enumerate(combined_df.iterrows()):
                current_row = start_row + index
                ws.cell(row=current_row, column=1, value=project_number)
                ws.cell(row=current_row, column=2, value=row_data.get('nome_arquivo', ''))
                ws.cell(row=current_row, column=3, value=row_data.get('qtd', 0))
                
                forma = str(row_data.get('forma', '')).lower()
                largura, altura = row_data.get('largura', 0), row_data.get('altura', 0)
                forma_map = {'circle': 'C', 'trapezoid': 'TP', 'right_triangle': 'T'}
                forma_abreviada = 'Q' if forma == 'rectangle' and largura == altura and largura > 0 else forma_map.get(forma, 'R' if forma == 'rectangle' else '')
                ws.cell(row=current_row, column=4, value=forma_abreviada)

                furos = row_data.get('furos', [])
                num_furos = len(furos) if isinstance(furos, list) else 0
                ws.cell(row=current_row, column=5, value=num_furos)
                ws.cell(row=current_row, column=6, value=furos[0].get('diam', 0) if num_furos > 0 else 0)
                ws.cell(row=current_row, column=7, value=row_data.get('espessura', 0))
                ws.cell(row=current_row, column=8, value=largura)
                ws.cell(row=current_row, column=9, value=altura)
                self.progress_bar.setValue(int(((index + 1) / (len(combined_df) * 2)) * 100))

            self.log_text.append("Calculando aproveitamento de chapas...")
            QApplication.processEvents()

            valid_nesting_df = combined_df[combined_df['forma'].isin(['rectangle', 'circle', 'right_triangle', 'trapezoid', 'dxf_shape'])].copy()
            valid_nesting_df['espessura'] = valid_nesting_df['espessura'].astype(float)
            grouped = valid_nesting_df.groupby('espessura')
            
            current_row = 209
            ws.cell(row=current_row, column=1, value="RELATÓRIO DE APROVEITAMENTO DE CHAPA").font = Font(bold=True, size=14)
            current_row += 2

            for espessura, group in grouped:
                # --- INÍCIO: LÓGICA DE OFFSET E MARGEM DINÂMICOS NA EXPORTAÇÃO EXCEL ---
                current_offset, current_margin = self._get_dynamic_offset_and_margin(espessura, offset, margin)
                # A margem efetiva é calculada para garantir 10mm da borda até a peça real.
                effective_margin = 10 - (current_offset / 2)
                # --- FIM: LÓGICA DE OFFSET E MARGEM ---
                pecas_para_calcular = []

                for _, row in group.iterrows():
                    # Adiciona peças à lista de cálculo, já com offset
                    # (A lógica para diferentes formas permanece a mesma)
                    if row['forma'] == 'rectangle' and row['largura'] > 0 and row['altura'] > 0:
                        pecas_para_calcular.append({'forma': 'rectangle', 'largura': row['largura'] + current_offset, 'altura': row['altura'] + current_offset, 'quantidade': int(row['qtd'])})
                    elif row['forma'] == 'circle' and row['diametro'] > 0:
                        pecas_para_calcular.append({'forma': 'circle', 'largura': row['diametro'] + current_offset, 'altura': row['diametro'] + current_offset, 'diametro': row['diametro'], 'quantidade': int(row['qtd'])})
                    elif row['forma'] == 'right_triangle' and row['rt_base'] > 0 and row['rt_height'] > 0:
                        pecas_para_calcular.append({'forma': 'right_triangle', 'largura': row['rt_base'] + current_offset, 'altura': row['rt_height'] + current_offset, 'quantidade': int(row['qtd'])})
                    elif row['forma'] == 'trapezoid' and row['trapezoid_large_base'] > 0 and row['trapezoid_height'] > 0:
                        pecas_para_calcular.append({'forma': 'trapezoid', 'largura': row['trapezoid_large_base'] + current_offset, 'altura': row['trapezoid_height'] + current_offset, 'small_base': row['trapezoid_small_base'] + current_offset, 'quantidade': int(row['qtd'])})
                    elif row['forma'] == 'dxf_shape' and row['largura'] > 0 and row['altura'] > 0:
                        pecas_para_calcular.append({'forma': 'dxf_shape', 'largura': row['largura'] + current_offset, 'altura': row['altura'] + current_offset, 'dxf_path': row['dxf_path'], 'quantidade': int(row['qtd'])})

                if not pecas_para_calcular: continue

                # --- CORREÇÃO: Garante que a função orquestradora seja chamada ---
                # A função orquestradora é essencial para a reutilização de sobras.
                # Ela executa o cálculo em duas fases para maximizar o aproveitamento.
                self.log_text.append(f"Otimizando espessura {espessura}mm (pode levar um momento)...")
                QApplication.processEvents()
                resultado = orquestrar_planos_de_corte(chapa_largura, chapa_altura, pecas_para_calcular, current_offset, effective_margin, espessura, status_signal_emitter=None)
                
                if not resultado: continue

                ws.cell(row=current_row, column=1, value=f"Espessura: {espessura} mm").font = Font(bold=True, size=12)
                current_row += 1
                total_chapas_usadas = resultado['total_chapas']
                peso_total_chapas_kg = (chapa_largura/1000) * (chapa_altura/1000) * espessura * 7.85 * total_chapas_usadas
                ws.cell(row=current_row, column=1, value=f"Total de Chapas: {total_chapas_usadas}")
                ws.cell(row=current_row, column=2, value=f"Aproveitamento: {resultado['aproveitamento_geral']}")
                ws.cell(row=current_row, column=3, value=f"Peso Total das Chapas: {peso_total_chapas_kg:.2f} kg").font = Font(bold=True)
                current_row += 2

                for i, plano_info in enumerate(resultado['planos_unicos']):
                    ws.cell(row=current_row, column=1, value=f"Plano de Corte {i+1} (Repetir {plano_info['repeticoes']}x)").font = Font(italic=True)
                    current_row += 1
                    ws.cell(row=current_row, column=2, value="Peças neste plano:")
                    current_row += 1
                    for item in plano_info['resumo_pecas']:
                        ws.cell(row=current_row, column=3, value=f"- {item['qtd']}x de {item['tipo']}")
                        current_row += 1
                    current_row += 1

                # --- INÍCIO: NOVA LÓGICA PARA ESCRITA DO RESUMO DE SUCATA ---
                sucata_info = resultado.get('sucata_detalhada')
                if sucata_info:
                    bold_font = Font(bold=True)
                    # 1. Peso do Offset
                    ws.cell(row=current_row, column=1, value="Peso do Offset (perda de corte):").font = bold_font
                    ws.cell(row=current_row, column=2, value=f"{sucata_info['peso_offset']:.2f} kg")
                    current_row += 2

                    # 2. Sobras Aproveitáveis
                    ws.cell(row=current_row, column=1, value="Sobras Aproveitáveis (Retalhos > 300x300 mm)").font = bold_font
                    current_row += 1
                    sobras_aproveitaveis = sucata_info['sobras_aproveitaveis']
                    if not sobras_aproveitaveis:
                        ws.cell(row=current_row, column=2, value="- Nenhuma")
                        current_row += 1
                    else:
                        from collections import Counter
                        contagem = Counter((s['largura'], s['altura'], f"{s['peso']:.2f}") for s in sobras_aproveitaveis for _ in range(s['quantidade']))
                        total_peso_aproveitavel = sum(s['peso'] * s['quantidade'] for s in sobras_aproveitaveis)
                        for (larg, alt, peso_unit), qtd in contagem.items():
                            ws.cell(row=current_row, column=2, value=f"- {qtd}x de {larg:.0f}x{alt:.0f} mm (Peso unit: {peso_unit} kg)")
                            current_row += 1
                        ws.cell(row=current_row, column=2, value=f"Peso Total Aproveitável: {total_peso_aproveitavel:.2f} kg").font = bold_font
                        current_row += 1
                    current_row += 1

                    # 3. Sucatas com Dimensões
                    ws.cell(row=current_row, column=1, value="Sucatas com Dimensões").font = bold_font
                    current_row += 1
                    sucatas_dim = sucata_info['sucatas_dimensionadas']
                    if not sucatas_dim:
                        ws.cell(row=current_row, column=2, value="- Nenhuma")
                        current_row += 1
                    else:
                        from collections import Counter
                        contagem = Counter((s['largura'], s['altura'], f"{s['peso']:.2f}") for s in sucatas_dim for _ in range(s['quantidade']))
                        total_peso_sucata_dim = sum(s['peso'] * s['quantidade'] for s in sucatas_dim)
                        for (larg, alt, peso_unit), qtd in contagem.items():
                            ws.cell(row=current_row, column=2, value=f"- {qtd}x de {larg:.0f}x{alt:.0f} mm (Peso unit: {peso_unit} kg)")
                            current_row += 1
                        ws.cell(row=current_row, column=2, value=f"Peso Total (Sucata Dimensionada): {total_peso_sucata_dim:.2f} kg").font = bold_font
                        current_row += 1
                    current_row += 1

                    # 4. Demais Sucatas
                    ws.cell(row=current_row, column=1, value="Demais Sucatas (cavacos, etc):").font = bold_font
                    ws.cell(row=current_row, column=2, value=f"{sucata_info['peso_demais_sucatas']:.2f} kg")
                    current_row += 2

                    # 5. Resumo de Perda Total
                    ws.cell(row=current_row, column=1, value="Resumo da Perda Total (Sucata + Processo + Offset):").font = bold_font
                    ws.cell(row=current_row, column=2, value=f"{resultado.get('peso_perda_total_sucata', 0):.2f} kg")
                    ws.cell(row=current_row, column=3, value=f"({resultado.get('percentual_perda_total_sucata', 0):.2f} % do total)").font = Font(italic=True)
                    current_row += 2
                # --- FIM: NOVA LÓGICA ---

                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
                cell = ws.cell(row=current_row, column=1)
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                current_row += 2
                self.progress_bar.setValue(50 + int((current_row / 400) * 50))

            self.log_text.append("Salvando arquivo Excel...")
            QApplication.processEvents()
            wb.save(save_path)
            self.progress_bar.setValue(100)
            self.log_text.append(f"Resumo do projeto salvo com sucesso em: {save_path}")
            QMessageBox.information(self, "Sucesso", f"O arquivo Excel foi salvo com sucesso em:\n{save_path}")
        except Exception as e:
            self.log_text.append(f"ERRO ao exportar para Excel: {e}")
            QMessageBox.critical(self, "Erro na Exportação", f"Ocorreu um erro ao salvar o arquivo:\n{e}")
        finally:
            self.set_buttons_enabled_on_process(True)
            self.progress_bar.setVisible(False)

    def _clear_session(self, clear_project_number=False):
        fields_to_clear = [self.nome_input, self.espessura_input, self.qtd_input, self.largura_input, self.altura_input, self.diametro_input, self.rt_base_input, self.rt_height_input, self.trapezoid_large_base_input, self.trapezoid_small_base_input, self.trapezoid_height_input, self.rep_diam_input, self.rep_offset_input, self.diametro_furo_input, self.pos_x_input, self.pos_y_input]
        if clear_project_number:
            fields_to_clear.append(self.projeto_input)
        for field in fields_to_clear:
            field.clear()
        self.furos_atuais = []
        self.update_furos_table()
        self.file_label.setText("Nenhum projeto ativo.")
        if clear_project_number: 
            self.excel_df = pd.DataFrame(columns=self.colunas_df)
            self.manual_df = pd.DataFrame(columns=self.colunas_df)
            self.update_table_display()

    def set_buttons_enabled_on_process(self, enabled):
        is_project_active = self.project_directory is not None
        has_items = not (self.excel_df.empty and self.manual_df.empty)
        self.calculate_nesting_btn.setEnabled(enabled and is_project_active and has_items)
        self.start_project_btn.setEnabled(enabled)
        self.history_btn.setEnabled(enabled)
        self.select_file_btn.setEnabled(enabled and is_project_active)
        self.import_dxf_btn.setEnabled(enabled and is_project_active) # <<< ATUALIZAÇÃO DE ESTADO >>>
        self.clear_excel_btn.setEnabled(enabled and is_project_active and not self.excel_df.empty)
        self.generate_code_btn.setEnabled(enabled and is_project_active)
        self.add_piece_btn.setEnabled(enabled and is_project_active)
        self.replicate_btn.setEnabled(enabled and is_project_active)
        self.add_furo_btn.setEnabled(enabled and is_project_active)
        self.process_pdf_btn.setEnabled(enabled and is_project_active and has_items)
        self.process_dxf_btn.setEnabled(enabled and is_project_active and has_items)
        self.process_all_btn.setEnabled(enabled and is_project_active and has_items)
        self.conclude_project_btn.setEnabled(enabled and is_project_active and has_items)
        self.export_excel_btn.setEnabled(enabled and is_project_active and has_items)

    def update_table_display(self):
        self.set_initial_button_state()
        # --- CORREÇÃO FUTUREWARNING: Concatena apenas os dataframes não vazios ---
        dfs_to_concat = [df for df in [self.excel_df, self.manual_df] if not df.empty]
        if dfs_to_concat:
            combined_df = pd.concat(dfs_to_concat, ignore_index=True)
        else:
            combined_df = pd.DataFrame(columns=self.colunas_df)
        self.pieces_table.blockSignals(True)
        self.pieces_table.setRowCount(0)
        self.pieces_table.blockSignals(False)

        if combined_df.empty:
            return

        self.pieces_table.setRowCount(len(combined_df))
        self.pieces_table.verticalHeader().setDefaultSectionSize(40)
        
        for i, row in combined_df.iterrows():
            for j, col in enumerate(self.colunas_df):
                value = row.get(col)
                if col == 'furos' and isinstance(value, list):
                    display_value = f"{len(value)} Furo(s)"
                elif pd.isna(value) or value == 0:
                    display_value = '-'
                else:
                    display_value = str(value)
                item = QTableWidgetItem(display_value)
                item.setTextAlignment(Qt.AlignVCenter | Qt.AlignLeft)
                self.pieces_table.setItem(i, j, item)

            action_widget = QWidget()
            action_layout = QHBoxLayout(action_widget)
            action_layout.setContentsMargins(5, 0, 5, 0)
            action_layout.setSpacing(5)
            edit_btn, delete_btn = QPushButton("Editar"), QPushButton("Excluir")
            edit_btn.clicked.connect(lambda _, r=i: self.edit_row(r))
            delete_btn.clicked.connect(lambda _, r=i: self.delete_row(r))
            action_layout.addWidget(edit_btn)
            action_layout.addWidget(delete_btn)
            self.pieces_table.setCellWidget(i, len(self.colunas_df), action_widget)

        header = self.pieces_table.horizontalHeader()
        header_map = {self.table_headers[i]: i for i in range(len(self.table_headers))}

        for col_name in ['Forma', 'Espessura', 'Qtd', 'Furos']:
            if col_name in header_map:
                header.setSectionResizeMode(header_map[col_name], QHeaderView.ResizeToContents)
        
        if 'Nome Arquivo' in header_map:
            header.setSectionResizeMode(header_map['Nome Arquivo'], QHeaderView.Stretch)
            
        dim_cols = ['Largura', 'Altura', 'Diametro', 'Rt Base', 'Rt Height', 
                    'Trapezoid Large Base', 'Trapezoid Small Base', 'Trapezoid Height']
        for col_name in dim_cols:
            if col_name in header_map:
                 header.setSectionResizeMode(header_map[col_name], QHeaderView.ResizeToContents)

        if 'Ações' in header_map:
            header.setSectionResizeMode(header_map['Ações'], QHeaderView.ResizeToContents)

    def edit_row(self, row_index):
        len_excel = len(self.excel_df)
        is_from_excel = row_index < len_excel
        df_source = self.excel_df if is_from_excel else self.manual_df
        local_index = row_index if is_from_excel else row_index - len_excel
        if local_index >= len(df_source): return # Proteção contra índice inválido
        piece_data = df_source.iloc[local_index]
        self.nome_input.setText(str(piece_data.get('nome_arquivo', '')))
        self.espessura_input.setText(str(piece_data.get('espessura', '')))
        self.qtd_input.setText(str(piece_data.get('qtd', '')))
        shape = piece_data.get('forma', '')
        index = self.forma_combo.findText(shape, Qt.MatchFixedString)
        if index >= 0: self.forma_combo.setCurrentIndex(index)
        self.largura_input.setText(str(piece_data.get('largura', '')))
        self.altura_input.setText(str(piece_data.get('altura', '')))
        self.diametro_input.setText(str(piece_data.get('diametro', '')))
        self.rt_base_input.setText(str(piece_data.get('rt_base', '')))
        self.rt_height_input.setText(str(piece_data.get('rt_height', '')))
        self.trapezoid_large_base_input.setText(str(piece_data.get('trapezoid_large_base', '')))
        self.trapezoid_small_base_input.setText(str(piece_data.get('trapezoid_small_base', '')))
        self.trapezoid_height_input.setText(str(piece_data.get('trapezoid_height', '')))
        self.furos_atuais = piece_data.get('furos', []).copy() if isinstance(piece_data.get('furos'), list) else []
        self.update_furos_table()
        df_source.drop(df_source.index[local_index], inplace=True)
        df_source.reset_index(drop=True, inplace=True)
        self.log_text.append(f"Peça '{piece_data['nome_arquivo']}' carregada para edição.")
        self.update_table_display()
    
    def delete_row(self, row_index):
        len_excel = len(self.excel_df)
        is_from_excel = row_index < len_excel
        df_source = self.excel_df if is_from_excel else self.manual_df
        local_index = row_index if is_from_excel else row_index - len_excel
        if local_index >= len(df_source): return # Proteção contra índice inválido
        piece_name = df_source.iloc[local_index]['nome_arquivo']
        df_source.drop(df_source.index[local_index], inplace=True)
        df_source.reset_index(drop=True, inplace=True)
        self.log_text.append(f"Peça '{piece_name}' removida.")
        self.update_table_display()
    
    def generate_piece_code(self):
        project_number = self.projeto_input.text().strip()
        if not project_number: QMessageBox.warning(self, "Campo Obrigatório", "Inicie um projeto para definir o 'Nº do Projeto'."); return
        new_code = self.code_generator.generate_new_code(project_number, prefix='VDS') #SUFIXO DOS CÓDIGOS
        if new_code: self.nome_input.setText(new_code); self.log_text.append(f"Código '{new_code}' gerado para o projeto '{project_number}'.")
    
    def add_manual_piece(self):
        try:
            nome = self.nome_input.text().strip()
            if not nome: QMessageBox.warning(self, "Campo Obrigatório", "'Nome/ID da Peça' é obrigatório."); return
            new_piece = {'furos': self.furos_atuais.copy()}
            for col in self.colunas_df:
                if col != 'furos': new_piece[col] = 0.0
            new_piece.update({'nome_arquivo': nome, 'forma': self.forma_combo.currentText()})
            fields_map = { 'espessura': self.espessura_input, 'qtd': self.qtd_input, 'largura': self.largura_input, 'altura': self.altura_input, 'diametro': self.diametro_input, 'rt_base': self.rt_base_input, 'rt_height': self.rt_height_input, 'trapezoid_large_base': self.trapezoid_large_base_input, 'trapezoid_small_base': self.trapezoid_small_base_input, 'trapezoid_height': self.trapezoid_height_input }
            for key, field in fields_map.items():
                new_piece[key] = float(field.text().replace(',', '.')) if field.text() else 0.0
            self.manual_df = pd.concat([self.manual_df, pd.DataFrame([new_piece])], ignore_index=True)
            self.log_text.append(f"Peça '{nome}' adicionada/atualizada.")
            self._clear_session(clear_project_number=False)
            self.update_table_display()
        except ValueError: QMessageBox.critical(self, "Erro de Valor", "Campos numéricos devem conter números válidos.")
    
    def select_file(self):
        if not self.project_directory:
            QMessageBox.warning(self, "Nenhum Projeto Ativo", "Inicie um projeto antes de carregar uma planilha.")
            return
        file_path, _ = QFileDialog.getOpenFileName(self, "Selecionar Planilha", "", "Excel Files (*.xlsx *.xls)")
        if file_path:
            try:
                df = pd.read_excel(file_path, header=0, decimal=','); df.columns = df.columns.str.strip().str.lower()
                df = df.loc[:, ~df.columns.duplicated()]
                for col in self.colunas_df:
                    if col not in df.columns: df[col] = pd.NA
                if 'furos' in df.columns:
                    def parse_furos(x):
                        if isinstance(x, list): return x
                        if isinstance(x, str) and x.startswith('['):
                            try: return json.loads(x.replace("'", "\""))
                            except json.JSONDecodeError: return []
                        return []
                    df['furos'] = df['furos'].apply(parse_furos)
                else:
                    df['furos'] = [[] for _ in range(len(df))]
                numeric_cols = [col for col in self.colunas_df if col != 'furos' and col != 'forma' and col != 'nome_arquivo']
                for col in numeric_cols:
                    if col in df.columns:
                        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
                self.excel_df = df[self.colunas_df]
                self.file_label.setText(f"Planilha: {os.path.basename(file_path)}"); self.update_table_display()
            except Exception as e: QMessageBox.critical(self, "Erro de Leitura", f"Falha ao ler o arquivo: {e}")
    
    def clear_excel_data(self):
        self.excel_df = pd.DataFrame(columns=self.colunas_df); self.file_label.setText("Nenhuma planilha selecionada"); self.update_table_display()

    def import_dxfs(self):
        if not self.project_directory:
            QMessageBox.warning(self, "Nenhum Projeto Ativo", "Inicie um projeto antes de importar arquivos DXF.")
            return

        file_paths, _ = QFileDialog.getOpenFileNames(self, "Selecionar Arquivos DXF", "", "DXF Files (*.dxf)")
        if not file_paths:
            return

        imported_count = 0
        for file_path in file_paths:
            largura, altura = get_dxf_bounding_box(file_path)

            if largura is not None and altura is not None:
                nome_arquivo = os.path.splitext(os.path.basename(file_path))[0]
                
                new_piece = { # type: ignore
                    'nome_arquivo': nome_arquivo,
                    'forma': 'rectangle', # Sempre será retângulo
                    'forma': 'dxf_shape',
                    'espessura': 0.0, # Padrão, para ser editado pelo usuário
                    'qtd': 1, # Padrão
                    'largura': round(largura, 2),
                    'altura': round(altura, 2),
                    'diametro': 0.0, 'rt_base': 0.0, 'rt_height': 0.0,
                    'trapezoid_large_base': 0.0, 'trapezoid_small_base': 0.0, 'trapezoid_height': 0.0,
                    'furos': [],
                    'dxf_path': file_path # Armazena o caminho do arquivo
                }
                self.manual_df = pd.concat([self.manual_df, pd.DataFrame([new_piece])], ignore_index=True)
                imported_count += 1
            else:
                self.log_text.append(f"AVISO: Não foi possível obter as dimensões do arquivo '{os.path.basename(file_path)}'. Pode estar vazio ou corrompido.")
        
        self.log_text.append(f"--- {imported_count} arquivo(s) DXF importado(s) com sucesso. ---")

    
    def replicate_holes(self):
        try:
            if self.forma_combo.currentText() != 'rectangle': QMessageBox.warning(self, "Função Indisponível", "Replicação disponível apenas para Retângulos."); return
            largura, altura = float(self.largura_input.text().replace(',', '.')), float(self.altura_input.text().replace(',', '.'))
            diam, offset = float(self.rep_diam_input.text().replace(',', '.')), float(self.rep_offset_input.text().replace(',', '.'))
            if (offset * 2) >= largura or (offset * 2) >= altura: QMessageBox.warning(self, "Offset Inválido", "Offset excede as dimensões da peça."); return
            furos = [{'diam': diam, 'x': offset, 'y': offset}, {'diam': diam, 'x': largura - offset, 'y': offset}, {'diam': diam, 'x': largura - offset, 'y': altura - offset}, {'diam': diam, 'x': offset, 'y': altura - offset}]
            self.furos_atuais.extend(furos); self.update_furos_table()
        except ValueError: QMessageBox.critical(self, "Erro de Valor", "Largura, Altura, Diâmetro e Offset devem ser números válidos.")
    
    def update_dimension_fields(self, shape):
        shape = shape.lower()
        is_rect, is_circ, is_tri, is_trap = shape == 'rectangle', shape == 'circle', shape == 'right_triangle', shape == 'trapezoid'
        for w in self.largura_row + self.altura_row: w.setVisible(is_rect)
        for w in self.diametro_row: w.setVisible(is_circ)
        for w in self.rt_base_row + self.rt_height_row: w.setVisible(is_tri)
        for w in self.trap_large_base_row + self.trap_small_base_row + self.trap_height_row: w.setVisible(is_trap)
        self.rep_group.setEnabled(is_rect)
    
    def add_furo_temp(self):
        try:
            diam, pos_x, pos_y = float(self.diametro_furo_input.text().replace(',', '.')), float(self.pos_x_input.text().replace(',', '.')), float(self.pos_y_input.text().replace(',', '.'))
            if diam <= 0: QMessageBox.warning(self, "Valor Inválido", "Diâmetro do furo deve ser maior que zero."); return
            self.furos_atuais.append({'diam': diam, 'x': pos_x, 'y': pos_y}); self.update_furos_table()
            for field in [self.diametro_furo_input, self.pos_x_input, self.pos_y_input]: field.clear()
        except ValueError: QMessageBox.critical(self, "Erro de Valor", "Campos de furo devem ser números válidos.")
    
    def update_furos_table(self):
        self.furos_table.setRowCount(0); self.furos_table.setRowCount(len(self.furos_atuais))
        for i, furo in enumerate(self.furos_atuais):
            self.furos_table.setItem(i, 0, QTableWidgetItem(str(furo['diam'])))
            self.furos_table.setItem(i, 1, QTableWidgetItem(str(furo['x'])))
            self.furos_table.setItem(i, 2, QTableWidgetItem(str(furo['y'])))
            delete_btn = QPushButton("Excluir")
            delete_btn.clicked.connect(lambda _, r=i: self.delete_furo_temp(r))
            self.furos_table.setCellWidget(i, 3, delete_btn)
        self.furos_table.resizeColumnsToContents()
    
    def delete_furo_temp(self, row_index):
        if 0 <= row_index < len(self.furos_atuais):
            del self.furos_atuais[row_index]
            self.update_furos_table()

# =============================================================================
# PONTO DE ENTRADA DA APLICAÇÃO
# =============================================================================
def main():
    app = QApplication(sys.argv)
    app.setStyleSheet(INOVA_PROCESS_STYLE)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())

if __name__ == "__main__":
    main()
    